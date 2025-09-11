from sqlalchemy.ext.asyncio import AsyncSession
from collections import defaultdict
from typing import Dict, NamedTuple, Optional, List
from datetime import datetime
from PIL import Image
from openpyxl.drawing.image import Image as OpenPyxlImage
import openpyxl
import textwrap
import os
from app.crud.export_p_chart import Export_P_Chart_CRUD
from app.utils.export_p_chart import Export_P_Chart_Utils
from collections import namedtuple
import math
from app.functions import parse_defect_string
import json
from dotenv import load_dotenv
import os
from app.functions import (
    get_days_in_month,
    transform_defect_data_to_defect_table,
    generate_month_data,
    sum_defects_by_day,
    calculate_defect_ratio,
    extract_fields_by_day,
)
import requests
import json

# from collections import defaultdict


class Export_P_Chart_Manager:
    def __init__(self):
        self.crud = Export_P_Chart_CRUD()
        self.utils = Export_P_Chart_Utils()
        self.BACKEND_API_SERVICE = os.environ.get("BACKEND_API_SERVICE")
        self.BACKEND_URL_SERVICE = os.environ.get("BACKEND_URL_SERVICE")

    def combine_json_list_defect_graph(self, data, skip_keys={}):
        Defect_Graph = namedtuple("Defect_Graph", ["id", "defect_name", "value"])
        safe_dict = {"Defect_Graph": Defect_Graph}
        data = eval(data, {}, safe_dict)
        result = {}
        # ---- 1. Combine defects by defect_name dynamically ----
        defect_dict = {}
        for entry in data:
            for defect in entry["defect"]:
                # Support both dict and object access
                if isinstance(defect, dict):
                    name = defect["defect_name"]
                    val = defect["value"]
                else:
                    name = getattr(defect, "defect_name")
                    val = getattr(defect, "value")
                if name not in defect_dict:
                    defect_dict[name] = list(val)
                else:
                    # sum elementwise
                    defect_dict[name] = [a + b for a, b in zip(defect_dict[name], val)]
        # convert back to list of dicts
        result["defect"] = [
            {"defect_name": name, "value": values}
            for name, values in defect_dict.items()
        ]
        # ---- 2. Combine other numeric lists dynamically ----
        # skip_keys = {"ucl1", "ucl2"}
        for key in data[0]:
            if key in ("defect", "p_bar"):  # skip 'defect' and 'p_bar' separately
                result[key] = data[0][key] if key != "defect" else result["defect"]
                continue
            if key in skip_keys:
                result[key] = data[0][key]
                continue
            # Only combine lists of numbers
            if isinstance(data[0][key], list) and all(
                isinstance(x, (int, float)) for x in data[0][key]
            ):
                # sum across all dicts for each position
                combined = [0] * len(data[0][key])
                for entry in data:
                    for i, val in enumerate(entry.get(key, [0] * len(combined))):
                        combined[i] += val
                result[key] = combined
            else:
                # Keep the first value (or implement special logic if needed)
                result[key] = data[0][key]
        # return str(result).replace("'", '"')
        return result

    def combine_json_list_defect_table(self, data, skip_keys={}):
        # Defect_table = namedtuple(
        #     "Defect_table",
        #     ["id", "defect_type", "defect_item", "category", "value"],
        #     defaults=(None,),
        # )
        class Defect_table(NamedTuple):
            id: int
            defect_type: str
            defect_item: str
            value: int
            category: Optional[List[str]] = None

        safe_dict = {"Defect_table": Defect_table}
        data = eval(data, {}, safe_dict)

        result = {}

        # Combine defect_table by defect_item
        sum_defect_tables = defaultdict(lambda: None)
        id_type_map = {}
        print("data:", data)
        # for entry in data:
        #     for d in entry["defect_table"]:
        #         name = d.defect_item
        #         values = d.value
        #         if sum_defect_tables[name] is None:
        #             sum_defect_tables[name] = list(values)
        #             id_type_map[name] = (d.id, d.defect_type)
        #         else:
        #             sum_defect_tables[name] = [
        #                 a + b for a, b in zip(sum_defect_tables[name], values)
        #             ]
        for entry in data:
            for d in entry["defect_table"]:
                key = (d.defect_item, d.defect_type)  # Use both as the unique key
                values = d.value
                if sum_defect_tables.get(key) is None:
                    sum_defect_tables[key] = list(values)
                    id_type_map[key] = d.id
                else:
                    sum_defect_tables[key] = [
                        a + b for a, b in zip(sum_defect_tables[key], values)
                    ]
        print("sum_defect_tables:", sum_defect_tables)
        print("id_type_map:", id_type_map)
        # Assemble the combined defect_table (as dicts; you can return as Defect_table if desired)
        # result["defect_table"] = [
        #     {
        #         "defect_item": name,
        #         "defect_type": id_type_map[name][1],
        #         "id": id_type_map[name][0],
        #         "value": vals,
        #     }
        #     for name, vals in sum_defect_tables.items()
        # ]
        result["defect_table"] = [
            {
                "defect_item": key[0],
                "defect_type": key[1],
                "id": id_type_map[key],
                "value": vals,
            }
            for key, vals in sum_defect_tables.items()
        ]
        print('result["defect_table"]:', result["defect_table"])
        # Combine all numerical lists by key (skip defect_table, keep day/index as per the first)
        for key in data[0]:
            if key == "defect_table":
                continue
            val0 = data[0][key]
            # Combine if numeric lists
            if isinstance(val0, list) and all(
                isinstance(x, (int, float)) for x in val0
            ):
                summed = [0] * len(val0)
                for entry in data:
                    for i, v in enumerate(entry.get(key, [0] * len(val0))):
                        summed[i] += v
                result[key] = summed
            else:
                result[key] = val0  # For "day", "index", etc. (string lists)
        for idx in range(len(result["defect_ratio"])):
            if result["prod_qty"][idx] != 0:
                result["defect_ratio"][idx] = round(
                    (result["defect_qty"][idx] / result["prod_qty"][idx] * 100), 2
                )
            else:
                result["defect_ratio"][idx] = 0
        # return str(result).replace("'", '"')
        return result

    async def fetch_pchart_defect_records_service(
        self, filters: Dict, db: AsyncSession = None
    ):
        """
        Service function to fetch records from `pchart_defect_record` table and generate an Excel report, then convert it to a PDF.

        Args:
            filters (Dict): Dictionary containing filter conditions.

        Returns:
            str: Path to the generated PDF file.
        """
        now = datetime.now()

        # Format as mm-yyyy
        month_year = now.strftime("%m-%Y")

        # Get the Unix timestamp
        unix_time = int(now.timestamp())
        current_directory = os.getcwd()
        if filters["process"] == "Outline":
            base_excel_path = (
                current_directory + "/app/utils/export_p_chart/Form_Outline.xlsx"
            )  # Base Excel file
        else:
            base_excel_path = (
                current_directory + "/app/utils/export_p_chart/Form.xlsx"
            )  # Base Excel file
        output_excel_path1 = (
            current_directory + "/app/utils/export_p_chart/p_chart_output1.xlsx"
        )  # Output file
        output_excel_path2 = (
            current_directory + "/app/utils/export_p_chart/p_chart_output2.xlsx"
        )  # Output file
        graph_image_path = (
            current_directory + "/app/utils/export_p_chart/temp_graph.png"
        )  # Path to the graph image
        output_pdf_path = (
            current_directory
            + f'/app/utils/export_p_chart/P-Chart-{filters [ "process" ]}-{filters [ "part_no" ]}-{month_year}-{unix_time}.pdf'
        )  # Final PDF file
        equation_image = current_directory + "/app/utils/export_p_chart/equation.png"
        equation_image2 = current_directory + "/app/utils/export_p_chart/equation2.png"

        records = await self.crud.fetch_filtered_records(db=db, filters=filters)
        for r in records:
            key_index = r._key_to_index
            part_name = (
                r[key_index["part_name"]] or filters["part_no"]
                if filters["process"] != "Outline"
                else "All"
            )
            n_bar = r[key_index["n_bar"]] or 0
            p_bar = r[key_index["p_bar"]] or 0
            k = r[key_index["k"]] or 0
            uclp = r[key_index["uclp"]] or 0
            lclp = r[key_index["lclp"]] or 0
            p_bar_last = r[key_index["p_bar_last"]] or 0

        pchart_graph = None
        #!
        graph_records = await self.crud.fetch_filtered_graph_records(
            db=db, filters=filters
        )
        # for r in graph_records:
        #     key_index = r._key_to_index
        #     pchart_graph = r[key_index["pchart_graph"]]
        result_graph = []
        for r in graph_records:
            key_index = r._key_to_index
            result_graph.append(r[key_index["pchart_graph"]])
        result_graph = "[" + ", ".join(str(item) for item in result_graph) + "]"
        # print("result_graph:", result_graph)
        pchart_graph = self.combine_json_list_defect_graph(
            result_graph,
            {"p_bar", "ucl_target", "x_axis_label", "x_axis_value", "y_right_axis"},
        )
        # print("pchart_graph:", pchart_graph)
        target = await self.crud.fetch_filtered_master_target_line(
            db=db, filters=filters
        )
        if target and len(target.fetchall()) > 0:
            for r in target:
                key_index = r._key_to_index
                target_control = r[key_index["target_control"]]
        else:
            target_control = "-"

        #!

        # for i in range(day_in_month + 1):
        #     print("i:", i)
        dt = datetime.strptime(filters["month"], "%B-%Y")
        date_str = dt.strftime("%Y-%m-01")
        day_in_month = get_days_in_month(filters["month"])
        list_prod_qty = [0] * day_in_month
        if filters["process"] == "Outline":
            line_id = self.crud.get_line_id(filters["line_name"])
            endpoint = (
                self.BACKEND_URL_SERVICE
                + "/api/prods/prod_qty?line_id="
                + str(line_id)
                + f"&shift={filters['shift']}"
                + f"&date={date_str}"
            )
            headers = {"X-API-Key": self.BACKEND_API_SERVICE}
            response_json = requests.get(endpoint, headers=headers).json()

            for i in range(0, len(response_json["prod_qty"])):
                c = int(str(response_json["prod_qty"][i]["production_date"])[8:10])
                list_prod_qty[c - 1] = response_json["prod_qty"][i]["actual_val"]

            list_prod_qty = list_prod_qty + [sum(list_prod_qty)]

            defect_outline = await self.crud.get_defect_outline(db=db, filters=filters)
            print("defect_outline:", defect_outline)
            defect_table_list = transform_defect_data_to_defect_table(
                defect_outline, day_in_month
            )
            day_table_list = generate_month_data(filters["month"])
            defect_qty_list = sum_defects_by_day(defect_outline, day_in_month)[
                "defect_qty"
            ]
            defect_ratio_list = calculate_defect_ratio(defect_qty_list, list_prod_qty)
            review_list = extract_fields_by_day(defect_outline)
            defect_outline_table = {
                "defect_table": defect_table_list,
                "day": day_table_list["day"],
                "index": day_table_list["index"],
                "prod_qty": list_prod_qty,
                "defect_qty": defect_qty_list,
                "defect_ratio": defect_ratio_list,
                "record_by": review_list["record_by"],
                "review_by_tl": review_list["review_by_tl"],
                "review_by_mgr": review_list["review_by_mgr"],
                "review_by_gm": review_list["review_by_gm"],
            }
            # print("defect_outline_table:", defect_outline_table)
            data_table = defect_outline_table
        else:
            table = await self.crud.fetch_filtered_table(db=db, filters=filters)
            result_table = []
            for r in table:
                key_index = r._key_to_index
                result_table.append(r[key_index["pchart_table"]])
            result_table = "[" + ", ".join(str(item) for item in result_table) + "]"
            # print("result_table:", result_table)
            table_pchart_table = self.combine_json_list_defect_table(result_table)
            # print("table_pchart_table1:", table_pchart_table)
            data_table = table_pchart_table
        # for r in table:
        #     key_index = r._key_to_index
        #     table_pchart_table = r[key_index["pchart_table"]]

        # print("type table_pchart_table1:", type(table_pchart_table))

        # table_pchart_table = self.utils.extract_pchart_table(table_pchart_table)
        # table_pchart_table = json.loads(table_pchart_table)
        # table_pchart_table = parse_defect_string(table_pchart_table)
        # print("-------------------------",table )
        # data_table = table_pchart_table
        # print("table_pchart_table2:", table_pchart_table)
        #!
        # date = []
        # trouble = []
        # action = []
        # in_change = []
        # manager = []

        # len_abnormal = 0

        # abnormal = await self.crud.fetch_filtered_abnormal(db=db, filters=filters)
        # for r in abnormal:
        #     key_index = r._key_to_index

        #     len_abnormal += 1
        #     date.append(r[key_index["date"]])
        #     trouble.append(r[key_index["trouble"]])
        #     action.append(r[key_index["action"]])
        #     in_change.append(r[key_index["in_change"]])
        #     manager.append(r[key_index["manager"]])
        #!
        # print(abnormal)
        # Load base Excel
        wb = openpyxl.load_workbook(base_excel_path)
        if "Page1" in wb.sheetnames:
            ws = wb["Page1"]
            ws2 = wb["Page2"]
        else:
            raise ValueError("Sheet 'Page1' not found in the Excel file.")

        # Process filters
        process_inline, process_outline, process_inspection = None, None, None
        if filters.get("process"):
            if filters["process"] == "Inline":
                process_inline = "✓"
            elif filters["process"] == "Outline":
                process_outline = "✓"
            elif filters["process"] == "Inspection":
                process_inspection = "✓"

        shift_a, shift_b = None, None
        if filters.get("shift"):
            if filters["shift"] == "A":
                shift_a = "✓"
            elif filters["shift"] == "B":
                shift_b = "✓"
            elif filters["shift"] == "All":
                shift_a, shift_b = "✓", "✓"

        # Define the data to insert into specific Excel cells
        month_cell = "AL3" if filters["process"] != "Outline" else "AC7"
        data_to_write = {
            month_cell: filters.get("month", ""),
            "R3": filters.get("line_name", ""),
            "R4": part_name,
            "R5": (
                (
                    filters.get("part_no", "")
                    + (
                        f' [{filters.get("sub_line_label", "")}]'
                        if filters.get("sub_line_label", "")
                        else ""
                    )
                )
                if filters["process"] != "Outline"
                else "All"
            ),
            "AC3": process_inline,
            "AC4": process_outline,
            "AF3": process_inspection,
            "AC5": shift_a,
            "AF5": shift_b,
        }
        if filters["process"] != "Outline":
            data_to_write.update(
                {
                    "AZ9": f"{n_bar:.2f}",
                    "AZ12": f"{p_bar:.2f}",
                    "AZ18": f"{k:.2f}",
                    "AZ22": f"{uclp:.2f}",
                    "AZ24": f"{lclp:.2f}",
                    "AY27": target_control,
                    "AY29": f"{p_bar_last:.5f}",
                }
            )
        start_col = 14
        if filters["process"] == "Outline":
            start_row = 8
        else:
            start_row = 30
        for i, day in enumerate(data_table["day"]):
            col_letter = self.utils.number_to_column_lower(start_col + i)
            row = str(start_row + 1)
            # cell_address = f"{col_letter}31"
            cell_address = f"{col_letter}{row}"
            data_to_write[cell_address] = day

        for i, day in enumerate(data_table["day"]):
            col_letter = self.utils.number_to_column_lower(start_col + i)
            row = str(start_row + 2)
            # cell_address = f"{col_letter}32"
            cell_address = f"{col_letter}{row}"
            data_to_write[cell_address] = i + 1

        for i, value in enumerate(data_table["prod_qty"]):
            col_letter = self.utils.number_to_column_lower(start_col + i)
            row = str(start_row + 3)
            # cell_address = f"{col_letter}33"
            cell_address = f"{col_letter}{row}"
            if i == len(data_table["prod_qty"]) - 1:
                # cell_address = f"AS33"
                cell_address = f"AS{row}"
            data_to_write[cell_address] = str(value)
            # addition
            if i == len(data_table["prod_qty"]) - 1:
                if filters["process"] != "Outline":
                    # cell_address = f"AV33"
                    cell_address = f"AV{row}"
                    data_to_write[cell_address] = str(value)

        for i, value in enumerate(data_table["defect_qty"]):
            col_letter = self.utils.number_to_column_lower(start_col + i)
            row = str(start_row + 4)
            # cell_address = f"{col_letter}34"
            cell_address = f"{col_letter}{row}"
            if i == len(data_table["defect_qty"]) - 1:
                # cell_address = f"AS34"
                cell_address = f"AS{row}"
            data_to_write[cell_address] = str(value)
            # addition
            if i == len(data_table["defect_qty"]) - 1:
                if filters["process"] != "Outline":
                    # cell_address = f"AV34"
                    cell_address = f"AV{row}"
                    data_to_write[cell_address] = str(value)

        # *******
        for i, value in enumerate(data_table["record_by"]):
            col_letter = self.utils.number_to_column_lower(start_col + i)
            row = str(start_row + 52)
            # cell_address = f"{col_letter}82"
            cell_address = f"{col_letter}{row}"
            # if i == len(data_table["defect_qty"]) - 1:
            #     cell_address = f"AS34"
            # print("value:", value)
            # print("type value:", type(value))
            data_to_write[cell_address] = str(value)
            # data_to_write[cell_address] = str(value[f'shift{filters [ "shift" ].lower()}'])
            # addition
            # if i == len(data_table["defect_qty"]) - 1:
            #     cell_address = f"AV34"
            # data_to_write[cell_address] = str(value)
        for i, value in enumerate(data_table["review_by_tl"]):
            col_letter = self.utils.number_to_column_lower(start_col + i)
            row = str(start_row + 53)
            # cell_address = f"{col_letter}83"
            cell_address = f"{col_letter}{row}"
            # if i == len(data_table["defect_qty"]) - 1:
            #     cell_address = f"AS34"
            # print("value:", value)
            # print("type value:", type(value))
            if type(value) is dict:
                data_to_write[cell_address] = str(
                    value[f'shift_{filters [ "shift" ].lower()}']
                )
            else:
                data_to_write[cell_address] = str(value)
            # addition
            # if i == len(data_table["defect_qty"]) - 1:
            #     cell_address = f"AV34"
            #     data_to_write[cell_address] = str(value)
        for i, value in enumerate(data_table["review_by_mgr"]):
            col_letter = self.utils.number_to_column_lower(start_col + i)
            row = str(start_row + 54)
            # cell_address = f"{col_letter}84"
            cell_address = f"{col_letter}{row}"
            # if i == len(data_table["defect_qty"]) - 1:
            #     cell_address = f"AS34"
            # data_to_write[cell_address] = str(value)
            if type(value) is dict:
                data_to_write[cell_address] = str(
                    value[f'shift_{filters [ "shift" ].lower()}']
                )
            else:
                data_to_write[cell_address] = str(value)
            # addition
            # if i == len(data_table["defect_qty"]) - 1:
            #     cell_address = f"AV34"
            #     data_to_write[cell_address] = str(value)
        for i, value in enumerate(data_table["review_by_gm"]):
            col_letter = self.utils.number_to_column_lower(start_col + i)
            row = str(start_row + 55)
            # cell_address = f"{col_letter}85"
            cell_address = f"{col_letter}{row}"
            # if i == len(data_table["defect_qty"]) - 1:
            #     cell_address = f"AS34"
            # data_to_write[cell_address] = str(value)
            if type(value) is dict:
                data_to_write[cell_address] = str(
                    value[f'shift_{filters [ "shift" ].lower()}']
                )
            else:
                data_to_write[cell_address] = str(value)
            # addition
            # if i == len(data_table["defect_qty"]) - 1:
            #     cell_address = f"AV34"
            #     data_to_write[cell_address] = str(value)
        # *****

        for i, value in enumerate(data_table["defect_ratio"]):
            col_letter = self.utils.number_to_column_lower(start_col + i)
            row = str(start_row + 5)
            # cell_address = f"{col_letter}35"
            cell_address = f"{col_letter}{row}"
            if i == len(data_table["defect_ratio"]) - 1:
                # cell_address = f"AS35"
                cell_address = f"AS{row}"
            data_to_write[cell_address] = str(value)
        # #!
        # sorted_defect_table = sorted(data_table["defect_table"], key=lambda x: x["id"])
        # sorted_defect_table = sorted_defect_table[:37]
        # counts = defaultdict(int)
        # # print("sorted_defect_table:", sorted_defect_table)
        # for row in sorted_defect_table:
        #     defect_type = row["defect_type"]
        #     counts[defect_type] += 1

        # # Convert defaultdict back to a regular dict (optional)
        # counts = dict(counts)
        # # print("counts:", counts)
        # row_del = 0
        # for i, value in enumerate(sorted_defect_table):
        #     initial_row = 36
        #     row = initial_row + i
        #     row_del = row + 1

        #     for j, value2 in enumerate(value["value"]):
        #         col_letter = self.utils.number_to_column_lower(start_col + j)
        #         if j == len(value["value"]) - 1:
        #             cell_address = f"AS{row}"
        #         else:
        #             cell_address = f"{col_letter}{row}"
        #         data_to_write[cell_address] = str(value2)

        #     if value["defect_item"] == "":
        #         value["defect_item"] = str(value["defect_type"])

        #     if value["defect_type"] not in ["M/C Set up", "Quality Test"]:
        #         cell_address = f"D{row}"
        #         data_to_write[cell_address] = str(value["id"])

        #     cell_address = f"F{row}"
        #     data_to_write[cell_address] = str(value["defect_item"])

        # # for header
        # initial_row = 35
        # merge_cell_list = []
        # # count_item=1
        # for key, value in counts.items():
        #     # selected_row = initial_row + max(int ( value/ 2 ),1)
        #     start = initial_row + 1
        #     initial_row = initial_row + value
        #     end = initial_row

        #     if key not in ["M/C Set up", "Quality Test"]:
        #         selected_row = start
        #         cell_address = f"C{selected_row}"
        #         data_to_write[cell_address] = str(key)
        #         if start < end:
        #             merge_cell_list.append(f"C{start}:C{end}")
        #             ws.merge_cells(f"C{start}:C{end}")

        # # Handle merged cells correctly
        # # print("data_to_write:", data_to_write)
        # for idx, (cell, value) in enumerate(data_to_write.items(), start=1):
        #     # print(f"idx:{idx} cell:{cell} value:{value}")
        #     # if idx <= 37:
        #     try:
        #         col_letter, row = openpyxl.utils.cell.coordinate_from_string(
        #             cell
        #         )  # Correctly unpack column and row
        #         col = openpyxl.utils.column_index_from_string(
        #             col_letter
        #         )  # Convert column letter to index
        #         ws.cell(row=row, column=col, value=value)
        #     except:
        #         pass
        # #!

        date = []
        trouble = []
        action = []
        in_change = []
        manager = []

        len_abnormal = 0
        if filters["process"] != "Outline":
            abnormal = await self.crud.fetch_filtered_abnormal(db=db, filters=filters)
            for r in abnormal:
                key_index = r._key_to_index

                len_abnormal += 1
                date.append(r[key_index["date"]])
                trouble.append(r[key_index["trouble"]])
                action.append(r[key_index["action"]])
                in_change.append(r[key_index["in_change"]])
                manager.append(r[key_index["manager"]])

            data_to_write_2 = {}

            MAX_ROW = 26
            LEFT_COL_OFFSET = 0  # columns B-G
            RIGHT_COL_OFFSET = 7  # columns I-N

            row = 5
            col_offset = LEFT_COL_OFFSET

            for i in range(len_abnormal):

                # ── wrap wide fields
                trouble_lines = textwrap.wrap(trouble[i], width=45)
                action_lines = textwrap.wrap(action[i], width=45)
                max_lines = max(len(trouble_lines), len(action_lines))

                # ── wrap narrow fields
                in_change_lines = textwrap.wrap(in_change[i], width=15)
                manager_lines = textwrap.wrap(manager[i], width=15)

                base_idx = 0
                remaining = max_lines

                while remaining:
                    free_lines = MAX_ROW - row + 1
                    lines_now = min(remaining, free_lines)

                    if base_idx == 0:
                        # write index and date only once (first line of record)
                        data_to_write_2[f"{chr ( 66 + col_offset )}{row}"] = i + 1
                        data_to_write_2[f"{chr ( 67 + col_offset )}{row}"] = date[i]

                        # write in_change and manager, only on first slice
                        for k in range(lines_now):
                            idx = base_idx + k
                            in_change_text = (
                                in_change_lines[idx]
                                if idx < len(in_change_lines)
                                else ""
                            )
                            manager_text = (
                                manager_lines[idx] if idx < len(manager_lines) else ""
                            )

                            data_to_write_2[f"{chr ( 70 + col_offset )}{row + k}"] = (
                                in_change_text
                            )
                            data_to_write_2[f"{chr ( 71 + col_offset )}{row + k}"] = (
                                manager_text
                            )
                    else:
                        # clear index/date in overflow rows
                        data_to_write_2[f"{chr ( 66 + col_offset )}{row}"] = ""
                        data_to_write_2[f"{chr ( 67 + col_offset )}{row}"] = ""

                        # clear in_change and manager in overflow rows
                        for k in range(lines_now):
                            data_to_write_2[f"{chr ( 70 + col_offset )}{row + k}"] = ""
                            data_to_write_2[f"{chr ( 71 + col_offset )}{row + k}"] = ""

                    # write trouble and action
                    for k in range(lines_now):
                        idx = base_idx + k
                        trouble_text = (
                            trouble_lines[idx] if idx < len(trouble_lines) else ""
                        )
                        action_text = (
                            action_lines[idx] if idx < len(action_lines) else ""
                        )

                        data_to_write_2[f"{chr ( 68 + col_offset )}{row + k}"] = (
                            trouble_text
                        )
                        data_to_write_2[f"{chr ( 69 + col_offset )}{row + k}"] = (
                            action_text
                        )

                    # advance
                    row += lines_now
                    base_idx += lines_now
                    remaining -= lines_now

                    # overflow → switch to right block
                    if remaining:
                        data_to_write_2[f"{chr ( 66 + col_offset )}{row}"] = ""
                        data_to_write_2[f"{chr ( 67 + col_offset )}{row}"] = ""
                        col_offset = RIGHT_COL_OFFSET
                        row = 5

            for cell, value in data_to_write_2.items():
                col_letter, row = openpyxl.utils.cell.coordinate_from_string(
                    cell
                )  # Correctly unpack column and row
                col = openpyxl.utils.column_index_from_string(
                    col_letter
                )  # Convert column letter to index
                ws2.cell(row=row, column=col, value=value)
        else:
            data_to_write_2 = {}

            MAX_ROW = 26
            LEFT_COL_OFFSET = 0  # columns B-G
            RIGHT_COL_OFFSET = 10  # columns I-N

            row = 5
            col_offset = LEFT_COL_OFFSET
            defect_outline_page2 = list(
                filter(
                    lambda d: (
                        int(d["defect_qty"]) > 0
                        if d["defect_qty"] not in ("", None, "None")
                        else False
                    ),
                    defect_outline,
                )
            )
            for i, defect in enumerate(defect_outline_page2):
                print("i:", i)
                print("defect:", defect)
                # ── wrap wide fields
                trouble_lines = textwrap.wrap(defect["master_defect_mode"], width=45)
                area_lines = textwrap.wrap(defect["master_defect_type"], width=45)
                partno_lines = textwrap.wrap(defect["part_no"], width=45)
                partname_lines = textwrap.wrap(defect["part_name"], width=45)
                qty_lines = textwrap.wrap(defect["defect_qty"], width=45)
                incharge_lines = textwrap.wrap(defect["pic"], width=45)
                max_lines = max(
                    len(trouble_lines),
                    len(area_lines),
                    len(partno_lines),
                    len(partname_lines),
                    len(qty_lines),
                )

                # # ── wrap narrow fields
                # in_change_lines = textwrap.wrap(in_change[i], width=15)
                # manager_lines = textwrap.wrap(manager[i], width=15)

                base_idx = 0
                remaining = max_lines

                while remaining:
                    free_lines = MAX_ROW - row + 1
                    lines_now = min(remaining, free_lines)

                    if base_idx == 0:
                        # write index and date only once (first line of record)
                        data_to_write_2[f"{chr ( 66 + col_offset )}{row}"] = i + 1
                        data_to_write_2[f"{chr ( 67 + col_offset )}{row}"] = defect[
                            "date"
                        ]

                        # write in_change and manager, only on first slice
                        # for k in range(lines_now):
                        #     idx = base_idx + k
                        #     in_change_text = (
                        #         in_change_lines[idx]
                        #         if idx < len(in_change_lines)
                        #         else ""
                        #     )
                        #     manager_text = (
                        #         manager_lines[idx] if idx < len(manager_lines) else ""
                        #     )

                        #     data_to_write_2[f"{chr ( 70 + col_offset )}{row + k}"] = (
                        #         in_change_text
                        #     )
                        #     data_to_write_2[f"{chr ( 71 + col_offset )}{row + k}"] = (
                        #         manager_text
                        #     )
                    else:
                        # clear index/date in overflow rows
                        data_to_write_2[f"{chr ( 66 + col_offset )}{row}"] = ""
                        data_to_write_2[f"{chr ( 67 + col_offset )}{row}"] = ""

                        # clear in_change and manager in overflow rows
                        for k in range(lines_now):
                            data_to_write_2[f"{chr ( 70 + col_offset )}{row + k}"] = ""
                            data_to_write_2[f"{chr ( 71 + col_offset )}{row + k}"] = ""

                    # write trouble and action
                    for k in range(lines_now):
                        idx = base_idx + k

                        partname_text = (
                            partname_lines[idx] if idx < len(partname_lines) else ""
                        )
                        partno_text = (
                            partno_lines[idx] if idx < len(partno_lines) else ""
                        )

                        area_text = area_lines[idx] if idx < len(area_lines) else ""
                        trouble_text = (
                            trouble_lines[idx] if idx < len(trouble_lines) else ""
                        )
                        qty_text = qty_lines[idx] if idx < len(qty_lines) else ""

                        incharge_text = (
                            incharge_lines[idx] if idx < len(incharge_lines) else ""
                        )

                        data_to_write_2[f"{chr ( 68 + col_offset )}{row + k}"] = (
                            partname_text
                        )
                        data_to_write_2[f"{chr ( 69 + col_offset )}{row + k}"] = (
                            partno_text
                        )
                        data_to_write_2[f"{chr ( 70 + col_offset )}{row + k}"] = (
                            area_text
                        )
                        data_to_write_2[f"{chr ( 71 + col_offset )}{row + k}"] = (
                            trouble_text
                        )
                        data_to_write_2[f"{chr ( 72 + col_offset )}{row + k}"] = (
                            qty_text
                        )
                        data_to_write_2[f"{chr ( 73 + col_offset )}{row + k}"] = "ST"
                        data_to_write_2[f"{chr ( 74 + col_offset )}{row + k}"] = (
                            incharge_text
                        )

                    # advance
                    row += lines_now
                    base_idx += lines_now
                    remaining -= lines_now

                    # overflow → switch to right block
                    if remaining:
                        data_to_write_2[f"{chr ( 66 + col_offset )}{row}"] = ""
                        data_to_write_2[f"{chr ( 67 + col_offset )}{row}"] = ""
                        col_offset = RIGHT_COL_OFFSET
                        row = 5

            for cell, value in data_to_write_2.items():
                col_letter, row = openpyxl.utils.cell.coordinate_from_string(
                    cell
                )  # Correctly unpack column and row
                col = openpyxl.utils.column_index_from_string(
                    col_letter
                )  # Convert column letter to index
                ws2.cell(row=row, column=col, value=value)

        # for idx, value in enumerate ( merge_cell_list ):
        #    ws.merge_cells ( value )
        # Process and insert the graph image
        # if "pchart_graph" in graph_records:.

        # Save the modified Excel file

        #!
        sorted_defect_table = sorted(data_table["defect_table"], key=lambda x: x["id"])
        print('data_table["defect_table"]:', data_table["defect_table"])
        # if filters['is_not_zero']:
        if filters["is_not_zero"]:
            # Step 1: Filter
            sorted_defect_table = [
                item for item in sorted_defect_table if sum(item["value"]) != 0
            ]

            # Step 2: Re-id
            re_ided = []
            for idx, item in enumerate(sorted_defect_table, start=1):
                new_item = item.copy()
                new_item["id"] = idx
                re_ided.append(new_item)
            sorted_defect_table = re_ided

        # print("sorted_defect_table:", sorted_defect_table)
        # print("sorted_defect_table.length", len(sorted_defect_table))
        defect_amount = len(sorted_defect_table)
        # print("page :", math.ceil(defect_amount / 37))
        page_amount = math.ceil(defect_amount / 46)
        ws2.title = f"Page{page_amount+2}"
        ws.title = f"Page{page_amount+3}"

        for idx_page in range(1, page_amount + 1):
            # print(f"Page{idx_page}")
            new_ws = wb.copy_worksheet(ws)
            new_ws.title = f"Page{idx_page}"
            # start_idx = (idx_page - 1) * 37
            # end_idx = (idx_page) * 37
            start_idx = (idx_page - 1) * 46
            end_idx = (idx_page) * 46
            if end_idx >= defect_amount:
                end_idx = defect_amount

            # print(f"data --> {start_idx}:{end_idx}")
            chunk_sorted_defect_table = sorted_defect_table[start_idx:end_idx]
            print("chunk_sorted_defect_table:", chunk_sorted_defect_table)
            # sorted_defect_table = sorted_defect_table[:37]
            ##!!

            counts = defaultdict(int)
            # print("sorted_defect_table:", sorted_defect_table)
            for row in chunk_sorted_defect_table:
                defect_type = row["defect_type"]
                counts[defect_type] += 1

            # Convert defaultdict back to a regular dict (optional)
            counts = dict(counts)
            # print("counts:", counts)
            row_del = 0
            for i, value in enumerate(chunk_sorted_defect_table):
                if filters["process"] == "Outline":
                    initial_row = 14
                else:
                    initial_row = 36
                row = initial_row + i
                row_del = row + 1

                for j, value2 in enumerate(value["value"]):
                    col_letter = self.utils.number_to_column_lower(start_col + j)
                    if j == len(value["value"]) - 1:
                        cell_address = f"AS{row}"
                    else:
                        cell_address = f"{col_letter}{row}"
                    data_to_write[cell_address] = str(value2)

                if value["defect_item"] == "":
                    value["defect_item"] = str(value["defect_type"])

                if value["defect_type"] not in ["M/C Set up", "Quality Test"]:
                    cell_address = f"D{row}"
                    data_to_write[cell_address] = str(value["id"])

                cell_address = f"F{row}"
                data_to_write[cell_address] = str(value["defect_item"])
            #!
            # for row in wb[f"Page{idx_page}"].iter_rows(
            #     min_row=row_del, max_row=initial_row + 45
            # ):
            #     for cell in row:
            #         cell.value = None
            # # for header
            if filters["process"] == "Outline":
                initial_row = 13
            else:
                initial_row = 35
            merge_cell_list = []
            # count_item=1
            for key, value in counts.items():
                # selected_row = initial_row + max(int ( value/ 2 ),1)
                start = initial_row + 1
                initial_row = initial_row + value
                end = initial_row

                if key not in ["M/C Set up", "Quality Test"]:
                    selected_row = start
                    cell_address = f"C{selected_row}"
                    data_to_write[cell_address] = str(key)
                    if start < end:
                        merge_cell_list.append(f"C{start}:C{end}")
                        wb[f"Page{idx_page}"].merge_cells(f"C{start}:C{end}")
                        # ws.merge_cells(f"C{start}:C{end}")

            # Handle merged cells correctly
            # print("data_to_write:", data_to_write)
            for idx, (cell, value) in enumerate(data_to_write.items(), start=1):
                # print(f"idx:{idx} cell:{cell} value:{value}")
                # if idx <= 37:
                try:
                    col_letter, row = openpyxl.utils.cell.coordinate_from_string(
                        cell
                    )  # Correctly unpack column and row
                    col = openpyxl.utils.column_index_from_string(
                        col_letter
                    )  # Convert column letter to index
                    wb[f"Page{idx_page}"].cell(row=row, column=col, value=value)
                    # ws.cell(row=row, column=col, value=value)
                except:
                    pass
            # wb[f"Page{idx_page}"].delete_rows(idx=row_del, amount=82 - row_del)
            print("pchart_graph:", pchart_graph)
            print("type pchart_graph:", type(pchart_graph))
            if pchart_graph != None:
                pchart_graph = self.utils.extract_pchart_graph(pchart_graph)

                if filters["is_not_zero"]:
                    # Filter out items with sum(value) == 0
                    filtered_list = [
                        item
                        for item in pchart_graph["defect_list"]
                        if sum(item["value"]) != 0
                    ]

                    # Resequence ids starting from 1
                    for idx, item in enumerate(filtered_list, start=1):
                        item["id"] = idx

                    # Update the original data
                    pchart_graph["defect_list"] = filtered_list

                # print("pchart_graph:", pchart_graph)
            if filters["process"] != "Outline":
                self.utils.create_graph_as_image(graph_image_path, pchart_graph)

                # Insert the graph image into a fixed position
                img = OpenPyxlImage(graph_image_path)
                img.anchor = "C8"
                wb[f"Page{idx_page}"].add_image(img)  # Adjust position as needed

                # ws.add_image(img)  # Adjust position as needed

                img = OpenPyxlImage(equation_image)
                img.anchor = "AT9"
                wb[f"Page{idx_page}"].add_image(img)  # Adjust position as needed
                # ws.add_image(img)  # Adjust position as needed

                img = OpenPyxlImage(equation_image2)
                img.anchor = "AT33"
                wb[f"Page{idx_page}"].add_image(img)  # Adjust position as needed
            # ws.add_image(img)  # Adjust position as needed
            # for row in wb[f"Page{idx_page}"].iter_rows(min_row=row_del, max_row=81):
            #     for cell in row:
            #         cell.value = None
            #!
            for row in wb[f"Page{idx_page}"].iter_rows(
                min_row=row_del, max_row=(81 if filters["process"] != "Outline" else 59)
            ):
                for cell in row:
                    cell.value = None
            #!wb[f"Page{idx_page}"].delete_rows(idx=row_del, amount=82 - row_del)
            # if filters["process"] != "Outline":
            #     wb[f"Page{idx_page}"].delete_rows(7, 23)
            # ws.delete_rows(idx=row_del, amount=82 - row_del)
        new_ws2 = wb.copy_worksheet(ws2)
        new_ws2.title = f"Page{page_amount+1}"
        wb.remove(wb[f"Page{page_amount+2}"])
        wb.remove(wb[f"Page{page_amount+3}"])
        #!
        wb.save(output_excel_path2)
        self.utils.convert_excel_to_pdf(output_excel_path2, output_pdf_path)
        # print("PDF report generated successfully.")
        if filters["file_type"] == "pdf":
            return output_pdf_path
        elif filters["file_type"] == "excel":
            return output_excel_path2
